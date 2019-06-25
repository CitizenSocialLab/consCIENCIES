import os

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")

import django
django.setup()

import settings

from openpyxl.reader.excel import load_workbook


def regenerate_translations(*args, **kwargs):
    """Import translations from XLSX file into MongoDB

        Optional kwargs:
        @param save: bool
        @param debug: bool

    """
    # Set params
    save = kwargs.get("save", False)
    debug = kwargs.get("debug", False)
    clean_db = kwargs.get("clean_db", False)
    filename = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "game/i18n/translations.xlsx") if "filename" not in kwargs else kwargs["filename"]


    # Load file
    #file = os.path.join()
    wb = load_workbook(filename=filename, use_iterators=True)



    # Loop through sheets (pages), collect definitions
    translations = {}
    for sheet_name in wb.get_sheet_names():
        # Init sheet data
        translations[sheet_name] = {}

        # Get the sheet
        sheet = wb.get_sheet_by_name(sheet_name)

        first_row = True
        for row in sheet.iter_rows():
            cells = [cell.internal_value for cell in row]




            # Construct dicts from language codes in header row
            if first_row:
                headers = cells[1:]
                if debug: print headers

                for lang_code in headers:
                    translations[sheet_name][lang_code] = {}

                first_row = False

            # Normal rows
            else:
                # Quit row loop if no key in first cell
                if not cells[0]: break

                # Field path from first cell in row, split by period
                field_pieces = cells[0].split('.')

                if debug: print "\n", field_pieces

                # Loop through rest of fields, mapping to lang code of column
                for c, cell in enumerate(cells[1:]):
                    lang_code = headers[c]

                    if debug: print lang_code

                    base = translations[sheet_name][lang_code]

                    # Loop through field pieces to assign value
                    for i, piece in enumerate(field_pieces):
                        if debug: print i, piece

                        if i+1 == len(field_pieces):
                            # Last field piece gets the value
                            base[piece] = cell
                            if debug: print base[piece]

                        else:
                            # Reset base to deeper level
                            if piece not in base:
                                base[piece] = {}

                            base = base[piece]

    # Save
    if save:
        print "\n\n"
        db = settings.MONGODB

        if clean_db:
            db.text.drop()

        for page, page_data in translations.iteritems():
            for lang, lang_data in page_data.iteritems():
                lang_data["page"] = page
                lang_data["lang"] = lang

                db.text.update({'page': page, 'lang': lang}, lang_data, upsert=True)

                print "updated %s %s" % (page, lang)


regenerate_translations(save=True, debug=False, clean_db=True)